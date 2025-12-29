from __future__ import annotations

import hashlib
from decimal import Decimal
from datetime import date
from typing import Iterable

from django.db import transaction
from openpyxl import load_workbook

from portfolios.models import Asset, Portfolio, Price, InitialHolding, DataImport


START_DATE_DEFAULT = date(2022, 2, 15)
V0_DEFAULT = Decimal("1000000000")


# -------------------------------------------------------------------
# Helpers
# -------------------------------------------------------------------

def file_sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def parse_decimal(value) -> Decimal:
    if value is None:
        return Decimal("0")

    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))

    s = str(value).strip()
    if s == "" or s in {"-", "—", "–", "N/A", "na", "null"}:
        return Decimal("0")

    s = s.replace(" ", "").replace(",", ".")
    if s.endswith("%"):
        return Decimal(s[:-1]) / Decimal("100")

    return Decimal(s)


# -------------------------------------------------------------------
# Excel readers
# -------------------------------------------------------------------

def read_weights(wb, start_date: date) -> tuple[dict[str, Decimal], dict[str, Decimal]]:
    """
    Hoja 'weights':
    Fecha | Activo | Portfolio 1 | Portfolio 2
    """
    ws = wb["weights"]

    w1, w2 = {}, {}

    for fecha, asset, p1, p2, *_ in ws.iter_rows(min_row=2, values_only=True):
        if not fecha or not asset:
            continue

        if hasattr(fecha, "date"):
            fecha = fecha.date()

        if fecha != start_date:
            continue

        code = str(asset).strip()
        w1[code] = parse_decimal(p1)
        w2[code] = parse_decimal(p2)

    if not w1 or not w2:
        raise ValueError(f"No weights encontrados para start_date={start_date}")

    return w1, w2


def read_prices(wb) -> Iterable[tuple[str, date, Decimal]]:
    """
    Hoja 'Precios':
    Dates | Asset1 | Asset2 | ...
    """
    ws = wb["Precios"]
    headers = [c.value for c in ws[1]]
    assets = [str(h).strip() for h in headers[1:] if h]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        dt = row[0].date() if hasattr(row[0], "date") else row[0]

        for idx, code in enumerate(assets, start=1):
            px = row[idx] if idx < len(row) else None
            if px is not None:
                yield code, dt, parse_decimal(px)


# -------------------------------------------------------------------
# ETL principal
# -------------------------------------------------------------------

def import_datos_xlsx(
    *,
    path: str,
    start_date: date = START_DATE_DEFAULT,
    v0: Decimal = V0_DEFAULT,
    force: bool = False,
) -> DataImport:
    """
    Flujo:
    1) Idempotencia por hash
    2) Lectura y normalización del Excel
    3) Persistencia incremental y creación del estado inicial
    """

    file_hash = file_sha256(path)

    if not force:
        existing = DataImport.objects.filter(file_hash=file_hash).first()
        if existing:
            return existing

    # --- Fase 1: lectura ---
    wb = load_workbook(filename=path, data_only=True)
    weights_1, weights_2 = read_weights(wb, start_date)
    price_rows = list(read_prices(wb))

    with transaction.atomic():

        data_import = DataImport.objects.create(
            source_name=path.split("/")[-1],
            file_hash=file_hash,
            status="SUCCESS",
        )

        # --- Fase 2: dominio base ---
        p1, _ = Portfolio.objects.get_or_create(
            name="Portfolio 1",
            defaults={"start_date": start_date, "initial_value": v0},
        )
        p2, _ = Portfolio.objects.get_or_create(
            name="Portfolio 2",
            defaults={"start_date": start_date, "initial_value": v0},
        )

        all_codes = set(weights_1) | set(weights_2) | {c for c, _, _ in price_rows}

        Asset.objects.bulk_create(
            [Asset(code=c, name=c) for c in all_codes],
            ignore_conflicts=True,
            batch_size=500,
        )

        assets = {a.code: a for a in Asset.objects.filter(code__in=all_codes)}

        # --- Fase 3: precios ---
        prices = [
            Price(asset=assets[c], date=dt, price=px)
            for c, dt, px in price_rows
            if c in assets
        ]

        created = Price.objects.bulk_create(
            prices, ignore_conflicts=True, batch_size=5000
        )

        data_import.rows_inserted = len(created)
        data_import.rows_updated = 0

        # --- Holdings iniciales ---
        if force:
            InitialHolding.objects.filter(portfolio__in=[p1, p2]).delete()

        prices_t0 = {
            p.asset_id: Decimal(p.price)
            for p in Price.objects.filter(date=start_date)
        }

        def create_holdings(portfolio, weights):
            rows = []
            for code, weight in weights.items():
                asset = assets.get(code)
                px0 = prices_t0.get(asset.id) if asset else None
                if px0:
                    qty = (weight * v0) / px0
                    rows.append(
                        InitialHolding(
                            portfolio=portfolio,
                            asset=asset,
                            quantity=qty,
                        )
                    )
            InitialHolding.objects.bulk_create(
                rows, ignore_conflicts=True, batch_size=500
            )

        create_holdings(p1, weights_1)
        create_holdings(p2, weights_2)

        data_import.save()

    return data_import
